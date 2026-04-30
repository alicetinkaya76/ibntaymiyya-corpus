"""H2 manuel doğrulama atölyesi için xlsx şablonu üretici.

Çıktı: docs/atolye/h2_review_template_v0.1.xlsx (4 sayfa)
    1. Talimatlar       — protokol, 5 etiket sınıfı tanımı, brief
    2. Etiketleme       — 26 pair, Hüseyin Hoca burada çalışır
    3. Claude Baseline  — gizli (atölye sırasında görünmemeli)
    4. Otomatik Özet    — sınıf dağılımı + bant × etiket çapraz tablo

Kullanım:
    python scripts/build_review_xlsx.py

Notlar:
- Sayfa 2 'Etiket' kolonu: data validation (5 sınıf dropdown)
- Sayfa 4: COUNTIF/COUNTIFS formülleri (Python'da hardcode değil)
- Claude baseline sayfası 'hidden' olarak işaretlenir; atölye sonrası
  Hüseyin Hoca + Ali tahkim için açılır.
"""

from __future__ import annotations

import json
import random
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_CANONICAL = REPO_ROOT / "data" / "canonical"
DATA_DERIVED = REPO_ROOT / "data" / "derived"
DATA_ANNOTATIONS = REPO_ROOT / "data" / "annotations" / "h2_review"
OUTPUT_DIR = REPO_ROOT / "docs" / "atolye"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

OUTPUT_PATH = OUTPUT_DIR / "h2_review_template_v0.1.xlsx"


# ---------------------------------------------------------------------------
# Veri toplama
# ---------------------------------------------------------------------------

def load_units(path: Path) -> dict:
    return {u["unit_id"]: u for u in json.loads(path.read_text(encoding="utf-8"))}


def stratified_sample() -> list[dict]:
    """Aynı seed ile aynı 26 pair (Streamlit ve script'le bit-perfect)."""
    pairs = json.loads(
        (DATA_DERIVED / "cross_corpus_h2_tfidf_char_v0.1.json").read_text()
    )["pairs"]
    top1 = {}
    for p in pairs:
        sid = p["source_unit_id"]
        if sid not in top1 or p["similarity"] > top1[sid]["similarity"]:
            top1[sid] = p
    top1_list = sorted(top1.values(), key=lambda p: -p["similarity"])

    rng = random.Random(42)
    bands = [
        (0.95, 1.01, 5, "Çok yüksek (likely birebir)"),
        (0.80, 0.95, 5, "Yüksek (büyük ihtimal düzenlenmiş tekrar)"),
        (0.60, 0.80, 10, "Orta-yüksek (paragraf ortak, paraphrase olabilir)"),
        (0.40, 0.60, 5, "Orta (parça ortak, asıl iddia farklı olabilir)"),
        (0.30, 0.40, 5, "Düşük (gürültü mü gerçek mi?)"),
    ]
    selected = []
    for lo, hi, n, label in bands:
        candidates = [p for p in top1_list if lo <= p["similarity"] < hi]
        sampled = rng.sample(candidates, min(n, len(candidates)))
        for p in sampled:
            p["band_label"] = label
        selected.extend(sampled)
    return selected


def load_claude_baseline() -> dict:
    """{(src, tgt): (label, notes)}"""
    files = sorted(DATA_ANNOTATIONS.glob("claude_4_7_baseline_*.json"))
    data = json.loads(files[-1].read_text(encoding="utf-8"))
    out = {}
    for a in data["annotations"]:
        out[(a["source_unit_id"], a["target_unit_id"])] = (a["label"], a["notes"])
    return out


# ---------------------------------------------------------------------------
# Stil sabitleri
# ---------------------------------------------------------------------------

# Renkler
HEADER_FILL = PatternFill("solid", start_color="2C3E50")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Arial")
SUBHEADER_FILL = PatternFill("solid", start_color="ECF0F1")

BAND_COLORS = {
    "Çok yüksek (likely birebir)": "C8E6C9",     # yeşil ton
    "Yüksek (büyük ihtimal düzenlenmiş tekrar)": "DCEDC8",
    "Orta-yüksek (paragraf ortak, paraphrase olabilir)": "FFF9C4",
    "Orta (parça ortak, asıl iddia farklı olabilir)": "FFE0B2",
    "Düşük (gürültü mü gerçek mi?)": "FFCCBC",
}

LABEL_COLORS = {
    "EXACT": "4CAF50",   # yeşil
    "EDITED": "8BC34A",  # açık yeşil
    "EXCERPT": "03A9F4", # mavi
    "OVERLAP": "FF9800", # turuncu
    "NOISE": "F44336",   # kırmızı
}

THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

ARABIC_FONT = Font(name="Times New Roman", size=12)
ARABIC_ALIGN = Alignment(
    horizontal="right", vertical="top", wrap_text=True, readingOrder=2
)
PROSE_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Sayfa 1 — Talimatlar
# ---------------------------------------------------------------------------

def build_instructions_sheet(wb):
    ws = wb.create_sheet("Talimatlar", 0)
    ws.sheet_view.showGridLines = False

    # Başlık
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "İbn Teymiyye Tutum Haritalama · H2 Manuel Doğrulama Atölyesi"
    c.font = Font(bold=True, size=16, name="Arial", color="2C3E50")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Açıklama
    ws.merge_cells("A2:C2")
    c = ws["A2"]
    c.value = (
        "Bu çalışma İbn Teymiyye'nin Iman bahsindeki paralel pasajların "
        "(Mecmû V07 İbn Kāsım edisyonu ↔ Standalone Iman Albani edisyonu) "
        "manuel doğrulamasıdır."
    )
    c.font = Font(italic=True, size=11, name="Arial", color="555555")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 36

    row = 4

    def write_section(title, body_lines, color="34495E"):
        nonlocal row
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = title
        c.font = Font(bold=True, size=13, name="Arial", color=color)
        c.alignment = Alignment(vertical="center")
        c.fill = SUBHEADER_FILL
        ws.row_dimensions[row].height = 24
        row += 1

        for line in body_lines:
            ws.merge_cells(f"A{row}:C{row}")
            c = ws[f"A{row}"]
            c.value = line
            c.font = Font(size=11, name="Arial")
            c.alignment = Alignment(vertical="top", wrap_text=True)
            # Yükseklik içerik uzunluğuna göre kaba ayarlama
            ws.row_dimensions[row].height = max(20, 18 * (1 + len(line) // 90))
            row += 1
        row += 1

    write_section("📋 PROTOKOL", [
        "1. Aşağıdaki 'Etiketleme' sekmesine geçin.",
        "2. Her satır bir pair (kaynak ↔ hedef pasaj). 26 pair var.",
        "3. Sırayla her pair için:",
        "    a. Sol kolonda 'Kaynak metin' (Mecmû V07 / İbn Kāsım edisyonu) okuyun.",
        "    b. Sağ kolonda 'Hedef metin' (Standalone Iman / Albani edisyonu) okuyun.",
        "    c. 'ETİKET' kolonuna tıklayın → açılır listeden 5 sınıftan birini seçin.",
        "    d. 'Notlar' kolonuna kararınızın gerekçesini yazın (özellikle OVERLAP / NOISE için).",
        "4. Excel her tuş vuruşunda otomatik kaydeder; özel bir 'kaydet' adımı gerekmiyor.",
        "5. 'Özet' sekmesinde ilerlemeniz canlı güncellenir (kaç tamamlandı, sınıf dağılımı).",
        "6. Tahmini süre: 90–120 dakika. Pair başı 3–5 dk.",
        "7. Bitince dosyayı (Sheets ise download xlsx, Excel ise .xlsx olduğu gibi) Ali'ye gönderin.",
    ])

    write_section("🏷️ 5 ETİKET SINIFI", [
        "EXACT — Birebir tekrar. Kelimelik farklar (tashkeel, noktalama, "
        "ayet referans bracket'ları) olabilir, ama metin akışı aynı.",
        "",
        "EDITED — Albani edisyonel müdahale yapmış. Fasıl başlığı çıkarılmış, "
        "öncül cümle atılmış, ya da paragraf sınırı yeniden çizilmiş; ama "
        "argüman yapısı aynı.",
        "",
        "EXCERPT — Hedef pasaj kaynağın bir kısmını içeriyor (özet veya alıntı). "
        "Yani Albani'nin pasajı Mecmû'nunkinden belirgin biçimde daha kısa.",
        "",
        "OVERLAP — Aynı geniş konuyu (iman, İslâm, küfür, münafık vs.) "
        "tartışıyor ama farklı argümanlar veya farklı pasajlar. Algoritma "
        "kelime ortaklığı yüzünden eşleştirmiş ama gerçek paralel değil.",
        "",
        "NOISE — Yanlış eşleşme. Konu bile uyuşmuyor. Algoritma yanılmış.",
    ])

    write_section("💡 ÖRNEKLER", [
        "EXACT örneği: İki pasaj harfi harfine aynı, sadece Albani parantez "
        "içinde [البقرة: 103] gibi ayet referansı eklemiş.",
        "",
        "EDITED örneği: Mecmû'da 'فصل: ومما يسأل عنه...' fasıl başlığıyla "
        "başlayan pasaj; Standalone'da bu başlık yok, 'والتحقيق...' kısmıyla "
        "başlıyor — sonrası birebir aynı.",
        "",
        "EXCERPT örneği: Mecmû'daki 800 kelimelik bir argüman, Standalone'da "
        "aynı argümanın özetlenmiş 200 kelimelik versiyonu olarak görünüyor.",
        "",
        "OVERLAP örneği: İkisi de Ahmad b. Hanbel'in iman hakkındaki rivayetlerini "
        "aktarıyor ama farklı rivayetler — ortak kelime kümesi yüksek "
        "(Ahmad, iman, İslâm) ama metinler farklı.",
        "",
        "NOISE örneği: Kaynak namaz bahsinden, hedef şefaat bahsinden; "
        "konuyla dahi alakası yok — sadece yüzeysel kelime ortaklığı.",
    ])

    write_section("📐 İPUÇLARI", [
        "• Şüpheliyseniz: kaynak ve hedef AYNI ARGÜMANI mı kuruyor sorusunu sorun. "
        "  Aynıysa EXACT/EDITED/EXCERPT'ten biri. Farklıysa OVERLAP veya NOISE.",
        "",
        "• EXACT vs EDITED ayrımı: cümle yapısı / kelime sırası değişmiş mi? "
        "  Hayırsa EXACT, evetse EDITED.",
        "",
        "• EDITED vs EXCERPT ayrımı: hedef pasajın UZUNLUĞU kaynağa göre belirgin "
        "  ölçüde daha kısa mı (en az %30 kayıp)? Evetse EXCERPT, hayırsa EDITED.",
        "",
        "• OVERLAP vs NOISE ayrımı: pasajlar AYNI GENİŞ KONUYU (iman, küfür vs.) mı "
        "  ele alıyor? Evetse OVERLAP, hayırsa NOISE.",
        "",
        "• Bant bilgisi (Çok yüksek / Yüksek / ...) algoritmanın ÖN TAHMİNİ. "
        "  Sizin etiketinize ipucu olabilir ama bağlayıcı DEĞİL — bant düşükken "
        "  EDITED, bant yüksekken OVERLAP çıkabilir. Metne güvenin.",
    ])

    write_section("ℹ️ TEKNİK NOTLAR", [
        "• Her pair için kaynak ve hedef metin yaklaşık 600–800 kelime. Tam "
        "  pasajlar verilmiştir; truncation yoktur.",
        "",
        "• 'Sayfalar' kolonu her pasajın hangi sayfa aralığında geçtiğini gösterir "
        "  (örn. V07P328–V07P417 → Mecmû Cilt 7'nin 328–417 sayfaları).",
        "",
        "• 'Skor' = TF-IDF kosinüs benzerlik (0–1). Yüksekse algoritma daha emin.",
        "  'Adjusted' = Kur'ân alıntıları çıkarıldıktan sonraki skor.",
        "  'QD-S' / 'QD-T' = kaynak / hedef'in Kur'ân yoğunluğu (0–1).",
        "",
        "• Kararsız kaldığınız pair'leri 'Notlar' kolonunda 'TARTIŞMA' yazarak "
        "  işaretleyin; atölye sonu birlikte tahkim edebiliriz.",
    ])

    # Kolon genişlikleri
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 35

    return ws


# ---------------------------------------------------------------------------
# Sayfa 2 — Etiketleme (Hüseyin Hoca'nın çalışacağı sayfa)
# ---------------------------------------------------------------------------

def build_annotation_sheet(wb, pairs, source_units, target_units):
    ws = wb.create_sheet("Etiketleme")
    ws.sheet_view.rightToLeft = False  # Tablo LTR; içerik hücreleri RTL

    headers = [
        "#",
        "Bant",
        "Source ID",
        "Target ID",
        "Skor",
        "Adjusted",
        "QD-S",
        "QD-T",
        "Sayfalar",
        "Kaynak metin (Mecmû V07 / İbn Kāsım)",
        "Hedef metin (Standalone Iman / Albani)",
        "ETİKET",
        "Notlar",
    ]

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

    ws.row_dimensions[1].height = 36

    # Veri satırları
    for i, pair in enumerate(pairs, start=1):
        row = i + 1

        src_unit = source_units.get(pair["source_unit_id"], {})
        tgt_unit = target_units.get(pair["target_unit_id"], {})

        page_s = (
            f"{src_unit['page_refs'][0]}–{src_unit['page_refs'][-1]}"
            if src_unit.get("page_refs") else "?"
        )
        page_t = (
            f"{tgt_unit['page_refs'][0]}–{tgt_unit['page_refs'][-1]}"
            if tgt_unit.get("page_refs") else "?"
        )

        values = [
            i,
            pair["band_label"],
            pair["source_unit_id"],
            pair["target_unit_id"],
            round(pair["similarity"], 3),
            round(pair["adjusted_similarity"], 3),
            round(pair["quran_density_source"], 2),
            round(pair["quran_density_target"], 2),
            f"S: {page_s}\nT: {page_t}",
            src_unit.get("full_text_arabic", ""),
            tgt_unit.get("full_text_arabic", ""),
            "",     # ETİKET ← Hüseyin Hoca dolduracak
            "",     # Notlar ← Hüseyin Hoca dolduracak
        ]

        for col_idx, val in enumerate(values, start=1):
            c = ws.cell(row=row, column=col_idx, value=val)
            c.border = THIN_BORDER
            c.font = Font(name="Arial", size=10)

            if col_idx in (10, 11):  # Arapça kolonlar
                c.font = ARABIC_FONT
                c.alignment = ARABIC_ALIGN
            elif col_idx == 13:  # Notlar
                c.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                c.alignment = CENTER_ALIGN

        # Bant rengi (B kolonu)
        band = pair["band_label"]
        if band in BAND_COLORS:
            ws.cell(row=row, column=2).fill = PatternFill(
                "solid", start_color=BAND_COLORS[band]
            )

        # Satır yüksekliği — Arapça metin için yeterli alan
        ws.row_dimensions[row].height = 280

    # Kolon genişlikleri
    widths = {
        "A": 4,    # #
        "B": 14,   # Bant
        "C": 18,   # Source ID
        "D": 16,   # Target ID
        "E": 7,    # Skor
        "F": 9,    # Adjusted
        "G": 7,    # QD-S
        "H": 7,    # QD-T
        "I": 14,   # Sayfalar
        "J": 65,   # Kaynak metin
        "K": 65,   # Hedef metin
        "L": 12,   # ETİKET
        "M": 35,   # Notlar
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Data validation: ETİKET kolonu (L) — 5 sınıf dropdown
    dv = DataValidation(
        type="list",
        formula1='"EXACT,EDITED,EXCERPT,OVERLAP,NOISE"',
        allow_blank=True,
        showDropDown=False,  # False = OK göster (tersine olduğunu unutma)
    )
    dv.error = "Sadece 5 sınıftan biri seçilebilir: EXACT/EDITED/EXCERPT/OVERLAP/NOISE"
    dv.errorTitle = "Geçersiz etiket"
    dv.prompt = "Açılır listeden bir etiket seçin"
    dv.promptTitle = "Etiket seçimi"
    ws.add_data_validation(dv)
    dv.add(f"L2:L{len(pairs) + 1}")

    # Conditional formatting: ETİKET kolonu rengi
    from openpyxl.formatting.rule import CellIsRule
    for label, color in LABEL_COLORS.items():
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{label}"'],
            fill=PatternFill("solid", start_color=color),
            font=Font(bold=True, color="FFFFFF", name="Arial"),
        )
        ws.conditional_formatting.add(f"L2:L{len(pairs) + 1}", rule)

    # Header satırını dondur
    ws.freeze_panes = "C2"

    return ws


# ---------------------------------------------------------------------------
# Sayfa 3 — Claude Baseline (atölye sırasında gizli)
# ---------------------------------------------------------------------------

def build_claude_baseline_sheet(wb, pairs, claude_baseline):
    ws = wb.create_sheet("Claude_Baseline_GIZLI")
    ws.sheet_state = "hidden"  # Atölye sırasında görünmesin

    headers = [
        "#", "Source ID", "Target ID", "Skor",
        "Claude'un etiketi", "Claude'un gerekçesi",
    ]

    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = PatternFill("solid", start_color="6A1B9A")  # mor — gizli
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    # Notice
    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = (
        "⚠️ Bu sayfa atölye SIRASINDA görünmemeli. Cohen κ (inter-rater "
        "agreement) hesaplanabilmesi için Hüseyin Hoca'nın bağımsız etiketleri "
        "öncelikli. Atölye SONRASI tahkim için açılır."
    )
    c.font = Font(italic=True, color="6A1B9A", size=10, name="Arial")
    c.fill = PatternFill("solid", start_color="F3E5F5")
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32

    for i, pair in enumerate(pairs, start=1):
        row = i + 2
        src = pair["source_unit_id"]
        tgt = pair["target_unit_id"]
        label, notes = claude_baseline.get((src, tgt), ("", ""))

        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=src)
        ws.cell(row=row, column=3, value=tgt)
        ws.cell(row=row, column=4, value=round(pair["similarity"], 3))
        ws.cell(row=row, column=5, value=label)
        ws.cell(row=row, column=6, value=notes)

        for col_idx in range(1, 7):
            c = ws.cell(row=row, column=col_idx)
            c.border = THIN_BORDER
            c.font = Font(name="Arial", size=10)
            if col_idx == 6:
                c.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                c.alignment = CENTER_ALIGN

        # Etiket rengi
        if label in LABEL_COLORS:
            c = ws.cell(row=row, column=5)
            c.fill = PatternFill("solid", start_color=LABEL_COLORS[label])
            c.font = Font(bold=True, color="FFFFFF", name="Arial")

        ws.row_dimensions[row].height = 80

    widths = {"A": 5, "B": 18, "C": 16, "D": 7, "E": 14, "F": 90}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# ---------------------------------------------------------------------------
# Sayfa 4 — Otomatik Özet (formüllerle)
# ---------------------------------------------------------------------------

def build_summary_sheet(wb, n_pairs):
    ws = wb.create_sheet("Özet")
    ws.sheet_view.showGridLines = False

    # Başlık
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "📊 İlerleme ve Sınıf Dağılımı"
    c.font = Font(bold=True, size=14, name="Arial", color="2C3E50")
    c.alignment = CENTER_ALIGN
    ws.row_dimensions[1].height = 28

    # İlerleme
    ws["A3"] = "Toplam pair:"
    ws["B3"] = n_pairs
    ws["A4"] = "Etiketlenen:"
    ws["B4"] = f'=COUNTA(Etiketleme!L2:L{n_pairs + 1})'
    ws["A5"] = "Kalan:"
    ws["B5"] = f"=B3-B4"
    ws["A6"] = "İlerleme:"
    ws["B6"] = f"=B4/B3"
    ws["B6"].number_format = "0.0%"

    for r in range(3, 7):
        ws.cell(row=r, column=1).font = Font(bold=True, name="Arial", size=11)
        ws.cell(row=r, column=2).font = Font(name="Arial", size=11)

    # Sınıf dağılımı
    ws["A9"] = "🏷️ Sınıf dağılımı"
    ws["A9"].font = Font(bold=True, size=12, name="Arial", color="2C3E50")
    ws["A9"].fill = SUBHEADER_FILL
    ws.merge_cells("A9:F9")

    ws["A10"] = "Etiket"
    ws["B10"] = "Sayı"
    ws["C10"] = "Oran"
    for col in (1, 2, 3):
        c = ws.cell(row=10, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    labels = ["EXACT", "EDITED", "EXCERPT", "OVERLAP", "NOISE"]
    for i, label in enumerate(labels):
        row = 11 + i
        ws.cell(row=row, column=1, value=label).font = Font(
            bold=True, color="FFFFFF", name="Arial",
        )
        ws.cell(row=row, column=1).fill = PatternFill(
            "solid", start_color=LABEL_COLORS[label]
        )
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN
        ws.cell(
            row=row, column=2,
            value=f'=COUNTIF(Etiketleme!L2:L{n_pairs + 1},"{label}")'
        )
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN
        # B6'da ilerleme oranı; burada toplam etiketli'ye bölelim
        ws.cell(
            row=row, column=3,
            value=f'=IF($B$4=0,0,B{row}/$B$4)'
        )
        ws.cell(row=row, column=3).number_format = "0.0%"
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

    # Bant × etiket çapraz tablo
    ws["A18"] = "📐 Bant × Etiket çapraz tablosu"
    ws["A18"].font = Font(bold=True, size=12, name="Arial", color="2C3E50")
    ws["A18"].fill = SUBHEADER_FILL
    ws.merge_cells("A18:F18")

    ws["A19"] = "Bant"
    for i, lbl in enumerate(labels):
        ws.cell(row=19, column=2 + i, value=lbl)

    for col in range(1, 7):
        c = ws.cell(row=19, column=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN

    bands = [
        ("Çok yüksek", "Çok yüksek (likely birebir)"),
        ("Yüksek", "Yüksek (büyük ihtimal düzenlenmiş tekrar)"),
        ("Orta-yüksek", "Orta-yüksek (paragraf ortak, paraphrase olabilir)"),
        ("Orta", "Orta (parça ortak, asıl iddia farklı olabilir)"),
        ("Düşük", "Düşük (gürültü mü gerçek mi?)"),
    ]

    for i, (short, full) in enumerate(bands):
        row = 20 + i
        ws.cell(row=row, column=1, value=short)
        ws.cell(row=row, column=1).font = Font(bold=True, name="Arial")
        ws.cell(row=row, column=1).fill = PatternFill(
            "solid", start_color=BAND_COLORS[full]
        )
        ws.cell(row=row, column=1).alignment = CENTER_ALIGN

        for j, lbl in enumerate(labels):
            col = 2 + j
            ws.cell(
                row=row, column=col,
                value=(
                    f'=COUNTIFS('
                    f'Etiketleme!B2:B{n_pairs + 1},"{full}",'
                    f'Etiketleme!L2:L{n_pairs + 1},"{lbl}")'
                ),
            )
            ws.cell(row=row, column=col).alignment = CENTER_ALIGN
            ws.cell(row=row, column=col).font = Font(name="Arial")

    # Özet ipucu
    ws["A27"] = "💡 Atölye Sonu"
    ws["A27"].font = Font(bold=True, size=12, name="Arial", color="2C3E50")
    ws.merge_cells("A27:F27")
    ws["A28"] = (
        "Ali, atölye sonu bu xlsx'i Claude'a gönderecek. Cohen κ "
        "(Hüseyin Hoca ↔ Claude baseline) hesaplanacak. Tartışmalı pair'ler "
        "(Notlar'da TARTIŞMA yazılı olanlar) Hüseyin Hoca + Ali ortak "
        "oturumda tahkim edilecek."
    )
    ws["A28"].font = Font(italic=True, size=10, name="Arial", color="555555")
    ws["A28"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A28:F28")
    ws.row_dimensions[28].height = 40

    # Kolon genişlikleri
    widths = {"A": 18, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("📥 Veri yükleniyor...")
    source_units = load_units(DATA_CANONICAL / "iman_v0.1.json")
    target_units = load_units(DATA_CANONICAL / "iman_standalone_v0.1.json")
    pairs = stratified_sample()
    claude = load_claude_baseline()
    print(f"  Source units: {len(source_units)}")
    print(f"  Target units: {len(target_units)}")
    print(f"  26 pair stratified: {len(pairs)}")
    print(f"  Claude baseline: {len(claude)} etiket")

    print("\n📝 Excel oluşturuluyor...")
    wb = Workbook()
    wb.remove(wb.active)  # default sheet

    build_instructions_sheet(wb)
    build_annotation_sheet(wb, pairs, source_units, target_units)
    build_claude_baseline_sheet(wb, pairs, claude)
    build_summary_sheet(wb, len(pairs))

    wb.save(OUTPUT_PATH)
    size_kb = OUTPUT_PATH.stat().st_size / 1024
    print(f"  Yazıldı: {OUTPUT_PATH} ({size_kb:.1f} KB)")
    print(f"\n✅ Sayfa sırası:")
    for ws in wb.worksheets:
        state = "🔒 GİZLİ" if ws.sheet_state == "hidden" else "👁️ görünür"
        print(f"  {state}  {ws.title}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
